import openpyxl
wb = openpyxl.load_workbook('C:/Users/aakash123/blender/reference/excel/YT_DataBase_12th _Class_Chapter(Physics).xlsx')

for s in ['1 Electric Charges and Fields', '5 Magnetism and Matter', '14 Semiconductor Electronics ,']:
    ws = wb[s]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"=== {s} ===")
    print("Headers:", headers)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        vals = list(row)
        eng = "yes" if (len(vals) > 2 and vals[2]) else "NO"
        hin = "yes" if (len(vals) > 3 and vals[3]) else "NO"
        print(f"  {vals[0]} | eng:{eng} | hin:{hin}")
    print()
