import openpyxl, json, re

wb = openpyxl.load_workbook('C:/Users/aakash123/blender/reference/excel/YT_DataBase_12th _Class_Chapter(Physics).xlsx')

# Class 12 Physics subject ID from DB
PHYSICS_SUBJECT_ID = 'cmmkpxm8d008puu342msauwlp'

# Read chapter list from '12th Chapters' sheet
ws = wb['12th Chapters']
chapters = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break
    chapters.append({
        'name': row[0],
        'chapterNo': int(row[1]),
        'sequence': int(row[2]),
        'subjectId': 1,
        'notesPdf': row[9] if len(row) > 9 and row[9] else None,
    })

print(f"Found {len(chapters)} chapters")

# Sheet name lookup by chapter number
sheet_lookup = {}
for s in wb.sheetnames:
    # Match "1 Electric..." or '"14 Semiconductor...'
    cleaned = s.strip().strip('"')
    m = re.match(r'^(\d+)\s', cleaned)
    if m:
        sheet_lookup[int(m.group(1))] = s

def extract_yt_id(url):
    if not url:
        return None
    m = re.search(r'youtu\.be/([a-zA-Z0-9_-]+)', str(url))
    if m:
        return m.group(1)
    m = re.search(r'[?&]v=([a-zA-Z0-9_-]+)', str(url))
    if m:
        return m.group(1)
    return None

data = []
for ch in chapters:
    sheet_name = sheet_lookup.get(ch['chapterNo'])
    if not sheet_name:
        print(f"  WARNING: No sheet for CH{ch['chapterNo']} {ch['name']}")
        data.append({**ch, 'dbSubjectId': PHYSICS_SUBJECT_ID, 'topics': []})
        continue

    ws2 = wb[sheet_name]
    topics = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if vals[0] is None:
            break
        # Headers: TopicName, IsExistE, VideoPathE, IsExistH, VideoPathH, TopicID, Thumbnail
        eng_id = extract_yt_id(vals[2]) if len(vals) > 2 else None
        hin_id = extract_yt_id(vals[4]) if len(vals) > 4 else None
        topics.append({
            'topicName': vals[0],
            'videoIdEnglish': eng_id,
            'videoIdHindi': hin_id,
        })

    data.append({
        **ch,
        'dbSubjectId': PHYSICS_SUBJECT_ID,
        'topics': topics,
    })

    has_eng = sum(1 for t in topics if t['videoIdEnglish'])
    has_hin = sum(1 for t in topics if t['videoIdHindi'])
    coming_soon = sum(1 for t in topics if not t['videoIdEnglish'] and not t['videoIdHindi'])
    print(f"  CH{ch['chapterNo']} {ch['name']}: {len(topics)} topics (eng:{has_eng}, hin:{has_hin}, coming_soon:{coming_soon})")

with open('C:/Users/aakash123/visuallearning-app/backend/import-data-12th.json', 'w') as f:
    json.dump(data, f, indent=2)

total_topics = sum(len(d['topics']) for d in data)
total_coming_soon = sum(1 for d in data for t in d['topics'] if not t['videoIdEnglish'] and not t['videoIdHindi'])
print(f"\nTotal: {len(data)} chapters, {total_topics} topics ({total_coming_soon} coming soon)")
